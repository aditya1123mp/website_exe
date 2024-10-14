from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time
import openpyxl
import selenium

def main():
# Initialize Chrome WebDriver
 chrome_options = Options()
 chrome_options.add_argument("--start-maximized")
 service = Service('C:\\Users\\DELL\\OneDrive\\Desktop\\chromedriver-win64\\chromedriver.exe')  # Replace with the path to your chromedriver
 driver = webdriver.Chrome(service=service, options=chrome_options)

# Open the App-Stick website
 driver.get('https://www.app-stick.com/')
 time.sleep(3)

# Define XPaths
 appstick_logo_xpath = "//div[@id='titlet']/a[text() = 'App-Stick']"
 sidebar_popular_xpath = "//div[@id='sidebar1']//a[text() ='Popular']"
 sidebar_programming_xpath = "//div[@id='sidebar1']//a[text() ='Programming']"

# Check if elements are present
 is_appstick_logo_present = len(driver.find_elements(By.XPATH, appstick_logo_xpath)) > 0
 is_sidebar_popular_present = len(driver.find_elements(By.XPATH, sidebar_popular_xpath)) > 0
 is_sidebar_programming_present = len(driver.find_elements(By.XPATH, sidebar_programming_xpath)) > 0

# Path to the Excel file
 file_path = "C:\\Users\\DELL\\OneDrive\\Desktop\\Websites.xlsx"

# Open the existing Excel file
 workbook = load_workbook(filename=file_path)
 sheet = workbook["Websites"]

# Find the next empty row
 row_count = sheet.max_row
 #data_row = sheet.row_dimensions[row_count + 1]  # Create a new row at the next available position

# Write 'open' if all elements are present, else write 'block'
 if is_appstick_logo_present and is_sidebar_popular_present and is_sidebar_programming_present:
  print("All elements are present, adding 'open' status to Excel file")
  sheet.append(["AppSticks", "https://www.app-stick.com/", "open"])
 else:
  print("One or more elements are not present, adding 'block' status to Excel file")
  sheet.append(["AppSticks", "https://www.app-stick.com/", "block"])

# Save the updated Excel file
 workbook.save(file_path)
 workbook.close()

 print(f"Excel file updated successfully at {file_path}")

# Close browser
 driver.quit()
# Execute the script only if it's run directly
if __name__ == "__main__":
    main()
