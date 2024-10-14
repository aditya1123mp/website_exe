from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time
import openpyxl
import selenium

def main():
# Path to the chromedriver
 chrome_driver_path = 'C:\\Users\\DELL\\OneDrive\\Desktop\\chromedriver-win64\\chromedriver.exe'

# Initialize Chrome WebDriver with options and Service
 chrome_options = Options()
 chrome_options.add_argument("--start-maximized")

# Use Service to set up the chromedriver path
 service = Service(executable_path=chrome_driver_path)
 driver = webdriver.Chrome(service=service, options=chrome_options)

# Navigate to the App-Stick website
 driver.get('https://www.liberkey.com/en.html')
 time.sleep(3)

# XPaths for the elements
 appsticklogoXPath = '//img[@id="logo"]'
 sidebarpolpular = "//a[@class='topdaddy']/span[text() = 'Applications']"
 sidebarprogramming = "//a[@class='topdaddy']/span[text() = 'Suites']"

# Check if the elements are present
 isAppStickLogoPresent = len(driver.find_elements(By.XPATH, appsticklogoXPath)) > 0
 isloginXPath = len(driver.find_elements(By.XPATH, sidebarpolpular)) > 0
 isdownloadXPath = len(driver.find_elements(By.XPATH, sidebarprogramming)) > 0

# Path to the existing Excel file
 file_path = "C:\\Users\\DELL\\OneDrive\\Desktop\\Websites.xlsx"

# Load the existing Excel file
 workbook = load_workbook(file_path)
 sheet = workbook.active

# Find the next empty row

 #rowCount = sheet.max_row

# If all elements are present, write 'open', otherwise write 'block'
 if isAppStickLogoPresent and isloginXPath and isdownloadXPath:
  print("All elements are present, adding 'open' status to Excel")
  sheet.append(["liberkey.com", "https://www.liberkey.com/en.html", "open"])
 else:
  print("One or more elements are not present, adding 'block' status to Excel")
  sheet.append(["liberkey.com", "https://www.liberkey.com/en.html", "block"])

# Save Excel file
 workbook.save(file_path)
 print(f"Excel file updated successfully at {file_path}")

# Close browser
 driver.quit()
# Execute the script only if it's run directly
if __name__ == "__main__":
    main()
